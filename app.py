"""
OrderFloz — Management Reporting Dashboard
==========================================
Vivant Skin Care | Wholesale B2B Sales Intelligence
"""

# =============================================================================
# IMPORTS
# =============================================================================
import streamlit as st
import pandas as pd
import requests
import pickle
import hashlib
import json
import io
import threading
from pathlib import Path
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED

# =============================================================================
# PAGE CONFIG
# =============================================================================
st.set_page_config(page_title="OrderFloz Reports", page_icon="📊", layout="wide")

# =============================================================================
# CONSTANTS
# =============================================================================
CONFIG_FILE     = Path(".orderfloz_reports_config.json")
CACHE_META_FILE = Path(".orderfloz_cache_meta.json")
CACHE_DIR       = Path(".orderfloz_cache")
CACHE_DIR.mkdir(exist_ok=True)

HUBSPOT_TTL_HOURS  = 24      # HubSpot tiers change infrequently
CONTACTS_TTL_HOURS = 12      # Cin7 contacts (reps/types) change infrequently
ORDERS_TTL_MINUTES = 15      # Open periods: re-check fingerprint every 15 min

# Cin7 limits: 500 rows/page is the max — use it
PAGE_SIZE   = 250
# Cin7 rate limit: 3 req/sec — keep batch small
PAGE_BATCH  = 3

CIN7_ORDER_FIELDS = (
    "id,company,billingCompany,firstName,lastName,"
    "email,total,createdDate,modifiedDate,orderDate,salesPersonId,source"
)

BRANDING = {"company_name": "OrderFloz", "primary_color": "#1a5276", "accent_color": "#00d4aa"}

# Approved account whitelist (committed to repo alongside app.py)
WHITELIST_FILE = Path("VIVANT CONTACT LIST FOR HUBSPOT.xlsx")

# =============================================================================
# CONFIG HELPERS
# =============================================================================

def load_config() -> dict:
    try:
        if CONFIG_FILE.exists():
            return json.loads(CONFIG_FILE.read_text())
    except Exception:
        pass
    return {}

def save_config(data: dict):
    try:
        existing = load_config()
        existing.update(data)
        CONFIG_FILE.write_text(json.dumps(existing))
    except Exception:
        pass

def get_secret(key: str, default: str = "") -> str:
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default

def get_excluded_domains() -> set:
    raw = load_config().get("excluded_domains", "") or get_secret("EXCLUDED_DOMAINS", "")
    if not raw:
        return set()
    return {d.strip().lower() for d in raw.split(",") if d.strip()}

# =============================================================================
# ACCOUNT WHITELIST  (loaded from Excel file committed to repo)
# =============================================================================

@st.cache_data(ttl=3600)
def load_account_whitelist() -> dict:
    """
    Load Active accounts from Excel. Key = Customer name (upper).
    Fields stored: tier (stripped of parens), rep (Owner Name, fallback to Rep initials).
    No st.* calls inside cached functions.
    """
    if not WHITELIST_FILE.exists():
        return {}
    try:
        import openpyxl, re as _re
        wb = openpyxl.load_workbook(WHITELIST_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        def col(name):
            try: return headers.index(name)
            except ValueError: return None

        idx_status   = col("Active Status")
        idx_customer = col("Customer")
        idx_tier     = col("Commission Tier")
        idx_rep      = col("Owner Name")
        idx_rep_init = col("Rep")

        accounts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if idx_status is not None and row[idx_status] != "Active":
                continue
            customer = (row[idx_customer] or "") if idx_customer is not None else ""
            name = str(customer).strip().upper()
            if not name:
                continue

            # Tier: strip surrounding parens — (HA) -> HA, (10%) -> 10%
            raw_tier = row[idx_tier] if idx_tier is not None else None
            tier = _re.sub(r"^\(|\)$", "", str(raw_tier).strip()).strip() if raw_tier else ""

            # Rep: use Owner Name; fall back to Rep initials if blank or is an email
            raw_owner = row[idx_rep]      if idx_rep      is not None else None
            raw_init  = row[idx_rep_init] if idx_rep_init is not None else None
            rep = (raw_owner or "").strip()
            if not rep or "@" in rep:
                rep = (raw_init or "").strip()

            accounts[name] = {"tier": tier, "rep": rep, "canonical": str(customer).strip()}

        return accounts
    except Exception:
        return {}
def _whitelist_lookup(company_upper: str, whitelist: dict) -> dict:
    """
    Fuzzy match a Cin7 company name against the whitelist.
    Returns entry dict with tier, rep, and canonical (the vetted Customer name).
    Cin7 names may have prefixes like '1 (FL) - ' or suffixes like ' (HA)'.
    """
    import re

    def _strip(s):
        s = re.sub(r'^\d+\s*\([A-Z]{2}\)\s*-\s*', '', s).strip()
        s = re.sub(r'\s*\((HA|6%|10%)\)\s*$', '', s).strip()
        return s

    # 1. Exact match
    if company_upper in whitelist:
        return whitelist[company_upper]

    # 2. Strip Cin7 prefix/suffix and try again
    cleaned = _strip(company_upper)
    if cleaned in whitelist:
        return whitelist[cleaned]

    # 3. Whitelist key fully contained in Cin7 name
    for key, val in whitelist.items():
        if len(key) >= 6 and key in company_upper:
            return val

    # 4. Cleaned Cin7 name contained in whitelist key (or vice versa)
    if len(cleaned) >= 6:
        for key, val in whitelist.items():
            if cleaned in key or key in cleaned:
                return val

    return {}

# =============================================================================
# DISK CACHE
# =============================================================================

def _cache_key(label: str) -> str:
    return hashlib.md5(label.encode()).hexdigest()

def _load_cache_meta() -> dict:
    try:
        if CACHE_META_FILE.exists():
            return json.loads(CACHE_META_FILE.read_text())
    except Exception:
        pass
    return {}

def _save_cache_meta(meta: dict):
    try:
        CACHE_META_FILE.write_text(json.dumps(meta))
    except Exception:
        pass

def cache_save_orders(label: str, orders: list, fingerprint: str):
    if not orders:
        return
    try:
        key  = _cache_key(label)
        path = CACHE_DIR / f"orders_{key}.pkl"
        with open(path, "wb") as f:
            pickle.dump(orders, f)
        meta = _load_cache_meta()
        meta[f"cin7_{key}"] = {
            "label": label, "fingerprint": fingerprint,
            "saved_at": datetime.now().isoformat(), "count": len(orders),
        }
        _save_cache_meta(meta)
    except Exception:
        pass

def cache_load_orders(label: str, fingerprint: str):
    try:
        key   = _cache_key(label)
        meta  = _load_cache_meta()
        entry = meta.get(f"cin7_{key}")
        if not entry or entry.get("count", 1) == 0:
            return None
        path = CACHE_DIR / f"orders_{key}.pkl"
        if not path.exists():
            return None
        if entry["fingerprint"] == fingerprint or fingerprint == "CLOSED":
            data = pickle.load(open(path, "rb"))
            return data if data else None
    except Exception:
        pass
    return None

def cache_has_orders(label: str) -> bool:
    """Check if ANY cached orders exist for label (no fingerprint check)."""
    try:
        key   = _cache_key(label)
        meta  = _load_cache_meta()
        entry = meta.get(f"cin7_{key}")
        if not entry or entry.get("count", 1) == 0:
            return False
        return (CACHE_DIR / f"orders_{key}.pkl").exists()
    except Exception:
        return False

def cache_load_orders_any(label: str):
    """Load cached orders regardless of fingerprint — for instant display."""
    try:
        key  = _cache_key(label)
        meta = _load_cache_meta()
        entry = meta.get(f"cin7_{key}")
        if not entry or entry.get("count", 1) == 0:
            return None
        path = CACHE_DIR / f"orders_{key}.pkl"
        if not path.exists():
            return None
        data = pickle.load(open(path, "rb"))
        return data if data else None
    except Exception:
        return None

def cache_save_hubspot(tiers: dict, owners: dict):
    try:
        path = CACHE_DIR / "hubspot.pkl"
        with open(path, "wb") as f:
            pickle.dump({"tiers": tiers, "owners": owners}, f)
        meta = _load_cache_meta()
        meta["hubspot"] = {"saved_at": datetime.now().isoformat()}
        _save_cache_meta(meta)
    except Exception:
        pass

def cache_load_hubspot():
    try:
        meta  = _load_cache_meta()
        entry = meta.get("hubspot")
        if not entry:
            return None, None
        age_h = (datetime.now() - datetime.fromisoformat(entry["saved_at"])).total_seconds() / 3600
        if age_h > HUBSPOT_TTL_HOURS:
            return None, None
        path = CACHE_DIR / "hubspot.pkl"
        if not path.exists():
            return None, None
        with open(path, "rb") as f:
            data = pickle.load(f)
        return data["tiers"], data["owners"]
    except Exception:
        pass
    return None, None

def cache_save_contacts(customers: dict):
    try:
        path = CACHE_DIR / "contacts.pkl"
        with open(path, "wb") as f:
            pickle.dump(customers, f)
        meta = _load_cache_meta()
        meta["contacts"] = {"saved_at": datetime.now().isoformat(), "count": len(customers)}
        _save_cache_meta(meta)
    except Exception:
        pass

def cache_load_contacts():
    try:
        meta  = _load_cache_meta()
        entry = meta.get("contacts")
        if not entry:
            return None
        age_h = (datetime.now() - datetime.fromisoformat(entry["saved_at"])).total_seconds() / 3600
        if age_h > CONTACTS_TTL_HOURS:
            return None
        path = CACHE_DIR / "contacts.pkl"
        if not path.exists():
            return None
        with open(path, "rb") as f:
            return pickle.load(f)
    except Exception:
        return None

def cache_clear_all():
    try:
        for f in CACHE_DIR.iterdir():
            f.unlink()
        if CACHE_META_FILE.exists():
            CACHE_META_FILE.unlink()
    except Exception:
        pass

# =============================================================================
# SESSION STATE
# =============================================================================

def _init_session():
    defaults = {
        "report_data":   None,
        "audit":         None,
        "periods":       [],
        "config_loaded": load_config(),
        "cin7_staff":    {},
        "cin7_customers":{},
        "fetching":      False,
        "fetch_status":  "",
        "last_fetch_ts": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_session()

# =============================================================================
# CIN7 — SINGLE PAGE FETCH (used internally by parallel fetcher)
# =============================================================================

def _fetch_page(username: str, api_key: str, start: str, end: str,
                page: int, use_fields: bool = True):
    """Fetch a single page. Returns (orders_list, hit_end).
    hit_end=True means this was the last page."""
    params = {
        "createdDateFrom": start,
        "createdDateTo":   end,
        "page":  page,
        "rows":  PAGE_SIZE,
    }
    if use_fields:
        params["fields"] = CIN7_ORDER_FIELDS
    try:
        r = requests.get(
            "https://api.cin7.com/api/v1/SalesOrders",
            auth=(username, api_key),
            params=params,
            timeout=60,
        )
        # Store debug info for first page
        if page == 1:
            try:
                with open(".cin7_api_debug.json", "w") as _f:
                    json.dump({
                        "status_code": r.status_code,
                        "url": r.url,
                        "params_sent": params,
                        "response_preview": r.text[:1000],
                    }, _f)
            except Exception:
                pass
        # Retry on 429 rate limit
        if r.status_code == 429:
            time.sleep(3)
            r = requests.get(
                "https://api.cin7.com/api/v1/SalesOrders",
                auth=(username, api_key),
                params=params,
                timeout=60,
            )
        # Retry without fields on 400/422
        if r.status_code in (400, 422) and use_fields:
            params.pop("fields", None)
            r = requests.get(
                "https://api.cin7.com/api/v1/SalesOrders",
                auth=(username, api_key),
                params=params,
                timeout=60,
            )
        if r.status_code != 200:
            return [], True
        orders = r.json() or []
        hit_end = len(orders) < PAGE_SIZE
        return orders, hit_end
    except Exception as ex:
        try:
            with open(".cin7_api_debug.json", "w") as _f:
                json.dump({"exception": str(ex)}, _f)
        except Exception:
            pass
        return [], True

# =============================================================================
# CIN7 — PARALLEL BATCH PAGE FETCH
# =============================================================================

import time

def fetch_orders_fast(username: str, api_key: str,
                      start_date: str, end_date: str,
                      label: str = "") -> list:
    """
    Fetch all orders for a date range using parallel page batching.
    Respects Cin7 rate limit: 3 req/sec, 60/min.
    """
    all_orders = []
    batch_start = 1

    while True:
        page_nums = range(batch_start, batch_start + PAGE_BATCH)

        with ThreadPoolExecutor(max_workers=PAGE_BATCH) as ex:
            future_map = {
                ex.submit(_fetch_page, username, api_key, start_date, end_date, p): p
                for p in page_nums
            }
            results = {}
            for f in as_completed(future_map):
                p = future_map[f]
                results[p] = f.result()

        # Process in page order; stop at first short page
        done = False
        for p in page_nums:
            orders, hit_end = results[p]
            all_orders.extend(orders)
            if hit_end:
                done = True
                break

        if done:
            break

        # Respect Cin7 rate limit: 3 req/sec (3 pages/batch, 1s gap = ~3 req/s max)
        time.sleep(1.0)
        batch_start += PAGE_BATCH

    return all_orders

# =============================================================================
# CIN7 — FINGERPRINT PROBE
# =============================================================================

def probe_cin7_fingerprint(username: str, api_key: str,
                            start_date: str, end_date: str) -> str:
    """Single-row probe to detect if data has changed."""
    try:
        r = requests.get(
            "https://api.cin7.com/api/v1/SalesOrders",
            auth=(username, api_key),
            params={
                "createdDateFrom": start_date,
                "createdDateTo":   end_date,
                "rows":   1,
                "order":  "modifiedDate desc",
                "fields": "id,modifiedDate",
            },
            timeout=10,
        )
        if r.status_code == 200:
            data = r.json()
            if data:
                o = data[0]
                return f"{o.get('id','')}:{o.get('modifiedDate','')}"
    except Exception:
        pass
    return ""

# =============================================================================
# CIN7 — STAFF
# =============================================================================

def fetch_cin7_staff(username: str, api_key: str) -> dict:
    if st.session_state.cin7_staff:
        return st.session_state.cin7_staff
    staff = {}
    try:
        r = requests.get(
            "https://api.cin7.com/api/v1/Users",
            auth=(username, api_key), timeout=15,
        )
        if r.status_code == 200:
            for u in r.json():
                uid  = u.get("id") or u.get("userId")
                name = f"{u.get('firstName','').strip()} {u.get('lastName','').strip()}".strip()
                if uid and name:
                    staff[uid] = name
    except Exception:
        pass
    st.session_state.cin7_staff = staff
    return staff

# =============================================================================
# CIN7 — CONTACTS  (disk-cached with TTL)
# =============================================================================

def fetch_cin7_customers(username: str, api_key: str) -> dict:
    """
    Fetch Cin7 Contacts. Returns {COMPANY_UPPER: {rep, type}}.
    Disk-cached with 12h TTL — contacts rarely change.
    """
    cached = cache_load_contacts()
    if cached is not None:
        return cached

    customers = {}
    page      = 1
    fields    = "id,name,salesRepresentative,customFields"

    while True:
        try:
            r = requests.get(
                "https://api.cin7.com/api/v1/Contacts",
                auth=(username, api_key),
                params={"page": page, "rows": 500, "fields": fields},
                timeout=30,
            )
            if r.status_code != 200:
                break
            data = r.json()
            if not data:
                break
            for c in data:
                name = (c.get("name") or "").strip().upper()
                if not name:
                    continue
                rep   = (c.get("salesRepresentative") or "").strip()
                cf    = c.get("customFields") or {}
                ctype = ""
                for k, v in cf.items():
                    if k == "Members_1037" or k.lower() == "type":
                        ctype = str(v).strip() if v else ""
                        break
                customers[name] = {"rep": rep, "type": ctype}
            if len(data) < 500:
                break
            page += 1
        except Exception:
            break

    cache_save_contacts(customers)
    return customers

# =============================================================================
# HUBSPOT
# =============================================================================

def test_cin7_connection(username: str, api_key: str) -> tuple:
    try:
        r = requests.get("https://api.cin7.com/api/v1/SalesOrders",
                         auth=(username, api_key), params={"rows": 1}, timeout=15)
        if r.status_code == 200: return True, "Connected"
        if r.status_code == 401: return False, "Invalid credentials"
        return False, f"HTTP {r.status_code}"
    except Exception as e:
        return False, str(e)

def test_hubspot_connection(api_key: str) -> tuple:
    try:
        r = requests.get("https://api.hubapi.com/crm/v3/objects/companies",
                         headers={"Authorization": f"Bearer {api_key}"},
                         params={"limit": 1}, timeout=15)
        if r.status_code == 200: return True, "Connected"
        if r.status_code == 401: return False, "Invalid API key"
        return False, f"HTTP {r.status_code}"
    except Exception as e:
        return False, str(e)

def fetch_hubspot_company_data(api_key: str, tier_property: str = "commission_tier") -> tuple:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    tiers = {}; owners = {}; after = None
    while True:
        params = {"limit": 100, "properties": f"name,{tier_property},hubspot_owner_id"}
        if after:
            params["after"] = after
        try:
            r = requests.get("https://api.hubapi.com/crm/v3/objects/companies",
                             headers=headers, params=params, timeout=30)
            if r.status_code != 200:
                break
            data = r.json()
            for company in data.get("results", []):
                props = company.get("properties", {})
                name  = (props.get("name") or "").strip().upper()
                if not name:
                    continue
                tiers[name]  = props.get(tier_property, "") or ""
                owner_id     = props.get("hubspot_owner_id") or ""
                if owner_id:
                    owners[name] = str(owner_id)
            after = data.get("paging", {}).get("next", {}).get("after")
            if not after:
                break
        except Exception:
            # ✅ FIX: removed st.warning() — not safe to call from threads
            break
    return tiers, owners

def fetch_hubspot_owners(api_key: str) -> dict:
    try:
        r = requests.get("https://api.hubapi.com/crm/v3/owners",
                         headers={"Authorization": f"Bearer {api_key}"},
                         params={"limit": 100}, timeout=15)
        if r.status_code != 200:
            return {}
        result = {}
        for o in r.json().get("results", []):
            oid  = str(o.get("id", ""))
            name = f"{o.get('firstName','').strip()} {o.get('lastName','').strip()}".strip() \
                   or o.get("email", "")
            if oid:
                result[oid] = name
        return result
    except Exception:
        return {}

# =============================================================================
# ORDER AGGREGATION
# =============================================================================

def aggregate_orders_by_company(orders: list, periods: list,
                                  cin7_staff: dict = None) -> tuple:
    excluded_domains = get_excluded_domains()
    period_labels    = [p["label"] for p in periods]

    def get_period_label(date_str):
        if not date_str:
            return None
        try:
            d = datetime.fromisoformat(date_str[:10]).date()
        except Exception:
            return None
        for p in periods:
            if p["start"] <= d <= p["end"]:
                return p["label"]
        return None

    def email_domain(email: str) -> str:
        email = (email or "").strip().lower()
        return email.split("@")[-1] if "@" in email else ""

    company_data = {}
    audit = {
        "total_raw": len(orders), "included": 0,
        "excluded_source": 0, "excluded_domain": 0,
        "excluded_no_period": 0, "excluded_zero_total": 0,
        "unknown_company": 0, "unique_companies": 0,
        "by_period": {lbl: {"included": 0, "excluded_source": 0, "revenue": 0.0}
                      for lbl in period_labels},
        "excluded_sources": {}, "excluded_domain_counts": {}, "sample_excluded": [],
    }

    for order in orders:
        company = (order.get("company") or order.get("billingCompany") or "").strip()
        if not company:
            first   = order.get("firstName", "")
            last    = order.get("lastName",  "")
            company = f"{first} {last}".strip() or "Unknown"
            if company == "Unknown":
                audit["unknown_company"] += 1

        total        = float(order.get("total") or 0)
        sp_id        = order.get("salesPersonId")
        rep_name     = (cin7_staff or {}).get(sp_id, "") if sp_id else ""
        cust_email   = (order.get("email") or "").strip()
        # Use orderDate for period matching — createdDate is system record date,
        # not the actual order/invoice date that Cin7 filters by
        order_date   = order.get("orderDate") or order.get("createdDate", "")
        source       = (order.get("source") or "").strip()
        source_lower = source.lower()
        period_label = get_period_label(order_date)

        if "shopify" in source_lower or "retail" in source_lower:
            audit["excluded_source"] += 1
            audit["excluded_sources"][source] = audit["excluded_sources"].get(source, 0) + 1
            if period_label in period_labels:
                audit["by_period"][period_label]["excluded_source"] += 1
            if len(audit["sample_excluded"]) < 10:
                audit["sample_excluded"].append({"reason": f"source={source}",
                    "company": company, "total": total,
                    "date": order_date[:10] if order_date else ""})
            continue

        domain = email_domain(cust_email)
        if excluded_domains and domain in excluded_domains:
            audit["excluded_domain"] += 1
            audit["excluded_domain_counts"][domain] = \
                audit["excluded_domain_counts"].get(domain, 0) + 1
            if len(audit["sample_excluded"]) < 10:
                audit["sample_excluded"].append({"reason": f"excluded domain ({domain})",
                    "company": company, "email": cust_email, "total": total,
                    "date": order_date[:10] if order_date else ""})
            continue

        if period_label is None or period_label not in period_labels:
            audit["excluded_no_period"] += 1
            continue

        if total == 0:
            audit["excluded_zero_total"] += 1

        if company not in company_data:
            company_data[company] = {"rep": rep_name, "order_count": 0}
            for lbl in period_labels:
                company_data[company][lbl] = 0.0

        company_data[company][period_label]  += total
        company_data[company]["order_count"] += 1
        if rep_name and not company_data[company]["rep"]:
            company_data[company]["rep"] = rep_name

        audit["included"]                            += 1
        audit["by_period"][period_label]["included"] += 1
        audit["by_period"][period_label]["revenue"]  += total

    audit["unique_companies"] = len(company_data)
    return company_data, audit

# =============================================================================
# REPORT DATAFRAME BUILDER
# =============================================================================

def build_report_dataframe(company_data: dict, hubspot_tiers: dict, periods: list,
                            company_owners: dict = None, owners_lookup: dict = None,
                            cin7_customers: dict = None) -> pd.DataFrame:
    labels        = [p["label"] for p in periods]
    primary_label = labels[-1]
    comp_label    = labels[-2] if len(labels) >= 2 else None
    primary_col   = primary_label
    comp_col      = comp_label if comp_label else "Comparison"

    # Load approved account whitelist
    whitelist = load_account_whitelist()
    use_whitelist = bool(whitelist)

    rows = []
    for company, data in company_data.items():
        primary_sales = data.get(primary_label, 0)
        comp_sales    = data.get(comp_label, 0) if comp_label else 0
        if primary_sales == 0 and comp_sales == 0:
            continue

        company_upper = company.upper()

        # Whitelist filter — skip accounts with no fuzzy match in approved list
        wl_entry = _whitelist_lookup(company_upper, whitelist) if use_whitelist else {}
        if use_whitelist and not wl_entry:
            continue

        # Tier from whitelist only
        tier  = wl_entry.get("tier", "")

        # Rep from whitelist first, fall back to cin7_customers and HubSpot owners
        rep = wl_entry.get("rep", "")
        if not rep:
            cin7_cust = (cin7_customers or {}).get(company_upper, {})
            rep = cin7_cust.get("rep", "") or data.get("rep", "")
        if not rep and company_owners and owners_lookup:
            owner_id = company_owners.get(company_upper, "")
            rep      = owners_lookup.get(owner_id, "") if owner_id else ""

        if comp_sales > 0:
            change_pct = ((primary_sales - comp_sales) / comp_sales) * 100
        elif primary_sales > 0:
            change_pct = 100.0
        else:
            change_pct = 0.0

        # Use the vetted Customer name from whitelist, not Cin7's raw name
        display_name = wl_entry.get("canonical") or company

        rows.append({
            "Account":         display_name,
            primary_col:       primary_sales,
            comp_col:          comp_sales,
            "$ Change":        primary_sales - comp_sales,
            "% Change":        change_pct,
            "Tier":            tier,
            "Sales Rep":       rep,
            "_order_count":    data.get("order_count", 0),
            "_primary_col":    primary_col,
            "_comparison_col": comp_col,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(primary_col, ascending=False).reset_index(drop=True)
    return df

# =============================================================================
# FULL FETCH PIPELINE  (called from button AND from auto-fetch)
# =============================================================================

def run_full_fetch(cin7_user: str, cin7_key: str, hubspot_key: str,
                   tier_property: str, periods: list) -> dict:
    """
    Execute the complete data fetch pipeline.
    Returns dict with keys: df, audit, hubspot_tiers, all_orders
    Designed to run fast via parallel batching.
    """
    today_dt   = datetime.now().date()
    all_orders = []

    # ── Step 1: For each period, decide probe or use CLOSED shortcut ──────────
    def _probe_or_skip(p, user, key):
        s  = p["start"].strftime("%Y-%m-%d")
        e  = p["end"].strftime("%Y-%m-%d")
        fp = "CLOSED" if p["end"] < today_dt else probe_cin7_fingerprint(user, key, s, e)
        return p, s, e, fp

    probed = []
    for p in periods:
        probed.append(_probe_or_skip(p, cin7_user, cin7_key))

    # ── Step 2: cache check ───────────────────────────────────────────────────
    # Also pull any preloaded monthly caches that overlap the requested periods
    # (avoids re-fetching when user preloaded data)
    import calendar as _cal

    def _overlapping_cached_months(periods_list):
        """Return orders from any cached month-label that overlaps a period."""
        meta = _load_cache_meta()
        loaded_labels = set()
        extra_orders = []
        for key, entry in meta.items():
            if not key.startswith("cin7_"):
                continue
            label = entry.get("label", "")
            if label in loaded_labels:
                continue
            # Parse "Mon YYYY" labels (preloaded months)
            try:
                month_start = datetime.strptime(label, "%b %Y").date()
                yr = month_start.year; mo = month_start.month
                month_end = month_start.replace(day=_cal.monthrange(yr, mo)[1])
            except ValueError:
                continue  # Not a monthly label
            # Check overlap with any requested period
            for p in periods_list:
                if month_start <= p["end"] and month_end >= p["start"]:
                    cached = cache_load_orders_any(label)
                    if cached:
                        extra_orders.extend(cached)
                        loaded_labels.add(label)
                    break
        return extra_orders

    needs_fetch = []
    for p, s, e, fp in probed:
        cached = cache_load_orders(p["label"], fp)
        if cached is not None:
            all_orders.extend(cached)
        else:
            needs_fetch.append((p, s, e, fp))

    # Load any preloaded monthly data that covers the gaps
    if needs_fetch:
        preloaded = _overlapping_cached_months(periods)
        if preloaded:
            all_orders.extend(preloaded)
            # Remove periods that are now covered by preloaded data
            covered = set()
            import calendar as _cal2
            for p, s, e, fp in needs_fetch:
                p_start = p["start"]; p_end = p["end"]
                # Check if all days in period are covered by preloaded orders
                all_dates = {o.get("orderDate","")[:7] for o in preloaded if o.get("orderDate")}
                needed_months = set()
                cur = p_start.replace(day=1)
                while cur <= p_end:
                    needed_months.add(cur.strftime("%Y-%m"))
                    m = cur.month + 1; y = cur.year + (1 if m > 12 else 0); m = m if m <= 12 else 1
                    cur = cur.replace(year=y, month=m, day=1)
                if needed_months.issubset(all_dates):
                    covered.add(p["label"])
            needs_fetch = [(p, s, e, fp) for p, s, e, fp in needs_fetch
                           if p["label"] not in covered]

    # ── Step 3: parallel fast-fetch all needed periods + HubSpot + Contacts ──
    hs_tiers  = {}
    hs_owners = {}
    hs_lookup = {}

    def _fetch_period(p, s, e, fp, user, key):
        orders = fetch_orders_fast(user, key, s, e, label=p["label"])
        cache_save_orders(p["label"], orders, fp)
        return orders

    # ✅ FIX: collect results into local variables — never touch st.* inside threads
    fetched_customers = None
    fetch_warnings    = []

    # ── Cin7 periods fetched SEQUENTIALLY to respect 3 req/sec rate limit ────
    for args in needs_fetch:
        try:
            orders = _fetch_period(*args, cin7_user, cin7_key)
            all_orders.extend(orders)
        except Exception as err:
            fetch_warnings.append(f"Fetch error (cin7): {err}")

    # ── HubSpot + Contacts in parallel (different API, no Cin7 rate limit) ───
    hs_tiers  = {}
    hs_owners = {}
    hs_lookup = {}

    def _fetch_hs(api_key, tier_prop):
        t, o = cache_load_hubspot()
        if t is not None:
            lk = fetch_hubspot_owners(api_key)
            return t, o, lk
        t, o = fetch_hubspot_company_data(api_key, tier_prop)
        lk   = fetch_hubspot_owners(api_key)
        cache_save_hubspot(t, o)
        return t, o, lk

    def _fetch_contacts(user, key):
        return fetch_cin7_customers(user, key)

    tasks = {}
    with ThreadPoolExecutor(max_workers=2) as ex:
        if hubspot_key:
            tasks[ex.submit(_fetch_hs, hubspot_key, tier_property)] = "hubspot"
        tasks[ex.submit(_fetch_contacts, cin7_user, cin7_key)] = "contacts"

        for fut in as_completed(tasks):
            kind = tasks[fut]
            try:
                if kind == "hubspot":
                    hs_tiers, hs_owners, hs_lookup = fut.result()
                elif kind == "contacts":
                    fetched_customers = fut.result()
            except Exception as err:
                fetch_warnings.append(f"Fetch error ({kind}): {err}")

    # ✅ FIX: apply to session_state and show warnings AFTER thread pool exits
    if fetched_customers is not None:
        st.session_state.cin7_customers = fetched_customers
    for w in fetch_warnings:
        st.warning(w)

    # ── Step 4: build DataFrame ───────────────────────────────────────────────
    cin7_staff   = fetch_cin7_staff(cin7_user, cin7_key)
    company_data, audit = aggregate_orders_by_company(
        all_orders, periods, cin7_staff=cin7_staff)
    df = build_report_dataframe(
        company_data, hs_tiers, periods, hs_owners, hs_lookup,
        cin7_customers=st.session_state.cin7_customers)

    # Store raw debug info to session state so it survives st.rerun()
    whitelist = load_account_whitelist()
    st.session_state["_debug"] = {
        "total_orders_fetched": len(all_orders),
        "periods": [p["label"] for p in periods],
        "first_order": {k: all_orders[0].get(k) for k in ["id","company","createdDate","orderDate","source","total"]} if all_orders else None,
        "audit_total_raw": audit.get("total_raw", 0),
        "audit_excluded_source": audit.get("excluded_source", 0),
        "audit_excluded_domain": audit.get("excluded_domain", 0),
        "audit_excluded_no_period": audit.get("excluded_no_period", 0),
        "audit_zero_total": audit.get("excluded_zero_total", 0),
        "source_values_seen": audit.get("excluded_sources", {}),
        "fetch_warnings": fetch_warnings,
        "whitelist_loaded": len(whitelist),
    }

    return {"df": df, "audit": audit, "hubspot_tiers": hs_tiers, "all_orders": all_orders}

# =============================================================================
# PRELOAD — BULK FETCH ALL MONTHS IN A DATE RANGE
# =============================================================================

def preload_months(cin7_user: str, cin7_key: str,
                   start_year: int, start_month: int,
                   end_year: int,   end_month: int,
                   progress_placeholder) -> dict:
    """
    Fetch and cache every calendar month between start and end (inclusive).
    Skips months already cached with a CLOSED fingerprint.
    Returns summary: {total_months, fetched, skipped, orders_loaded}
    """
    import calendar as _cal

    today = datetime.now().date()

    # Build list of (year, month) tuples
    months = []
    y, m = start_year, start_month
    while (y, m) <= (end_year, end_month):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1

    summary = {"total": len(months), "fetched": 0, "skipped": 0, "orders": 0}

    for i, (y, m) in enumerate(months):
        ld    = _cal.monthrange(y, m)[1]
        start = datetime(y, m, 1).date()
        end   = min(datetime(y, m, ld).date(), today)
        label = datetime(y, m, 1).strftime("%b %Y")
        s_str = start.strftime("%Y-%m-%d")
        e_str = end.strftime("%Y-%m-%d")

        pct = i / len(months)
        progress_placeholder.progress(pct, text=f"📦 {label} ({i+1}/{len(months)})…")

        # Closed months: use fingerprint "CLOSED" — no probe needed
        is_closed = end < today
        fp = "CLOSED" if is_closed else probe_cin7_fingerprint(cin7_user, cin7_key, s_str, e_str)

        # Skip if already cached with matching fingerprint
        cached = cache_load_orders(label, fp)
        if cached is not None:
            summary["skipped"] += 1
            summary["orders"]  += len(cached)
            continue

        # Fetch the month
        orders = fetch_orders_fast(cin7_user, cin7_key, s_str, e_str, label=label)
        cache_save_orders(label, orders, fp)
        summary["fetched"] += 1
        summary["orders"]  += len(orders)

        # Small gap between months (probe already consumed 1 req)
        if not is_closed:
            time.sleep(1.0)

    progress_placeholder.progress(1.0, text=f"✅ Done — {summary['orders']:,} orders across {len(months)} months")
    return summary


# =============================================================================
# EXPORT
# =============================================================================

def export_to_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sales Report", index=False)
        ws = writer.sheets["Sales Report"]
        for idx, col in enumerate(df.columns):
            width = min(max(df[col].astype(str).map(len).max(), len(col)) + 2, 50)
            ws.column_dimensions[chr(65 + idx)].width = width
    return output.getvalue()

def export_raw_orders_csv() -> bytes:
    """
    Load all cached orders and export as CSV, sorted by order date descending.
    Adds order_date (clean YYYY-MM-DD) and order_year_month (YYYY-MM) for easy sorting/filtering.
    """
    meta = _load_cache_meta()
    all_orders = []
    for key, entry in meta.items():
        if not key.startswith("cin7_"):
            continue
        cached = cache_load_orders_any(entry.get("label", ""))
        if cached:
            all_orders.extend(cached)

    if not all_orders:
        return b""

    # Deduplicate by order id
    seen = set()
    unique_orders = []
    for o in all_orders:
        oid = o.get("id")
        if oid and oid in seen:
            continue
        if oid:
            seen.add(oid)
        unique_orders.append(o)

    # Load whitelist for account/tier/rep enrichment (guard against None)
    wl = load_account_whitelist() or {}

    rows = []
    for o in unique_orders:
        row = {k: v for k, v in o.items() if not isinstance(v, (dict, list))}

        # Clean date columns
        raw_date = o.get("orderDate") or o.get("createdDate", "")
        try:
            od = datetime.strptime(raw_date[:10], "%Y-%m-%d").date()
            row["order_date"]       = od.strftime("%Y-%m-%d")
            row["order_year_month"] = od.strftime("%Y-%m")
        except:
            row["order_date"]       = ""
            row["order_year_month"] = ""

        # Whitelist enrichment
        company  = o.get("company") or o.get("billingCompany") or ""
        wl_entry = _whitelist_lookup(company.upper(), wl)
        if wl_entry:
            row["account"] = wl_entry.get("canonical", company)
            row["tier"]    = wl_entry.get("tier", "")
            row["rep"]     = wl_entry.get("rep", "")
        else:
            row["account"] = company
            row["tier"]    = ""
            row["rep"]     = ""

        rows.append(row)

    df = pd.DataFrame(rows)
    if "order_date" in df.columns:
        df = df.sort_values("order_date", ascending=False)

    priority = [
        "order_date", "order_year_month",
        "account", "tier", "rep",
        "id", "reference", "company", "billingCompany",
        "firstName", "lastName", "email",
        "total", "source", "status", "salesPersonId",
    ]
    cols = [c for c in priority if c in df.columns]
    rest = [c for c in df.columns if c not in cols]
    df   = df[cols + rest]

    return df.to_csv(index=False).encode("utf-8")


# =============================================================================
# CSS
# =============================================================================

def inject_css():
    st.markdown(f"""
    <style>
        .main-header {{ text-align:center; padding:0.75rem 0 0.5rem; }}
        .main-header h1 {{ color:{BRANDING['primary_color']}; margin:0; font-size:1.8rem; }}
        .stale-banner {{
            background:#fff3cd; border:1px solid #ffc107; border-radius:6px;
            padding:0.4rem 0.8rem; font-size:0.85rem; margin-bottom:0.5rem;
        }}
    </style>
    """, unsafe_allow_html=True)

# =============================================================================
# CHARTS
# =============================================================================

def create_yoy_chart(df):
    import plotly.graph_objects as go
    pc = df["_primary_col"].iloc[0]; cc = df["_comparison_col"].iloc[0]
    top = df.nlargest(15, pc)
    fig = go.Figure()
    if cc in df.columns:
        fig.add_trace(go.Bar(name=cc, x=top["Account"], y=top[cc], marker_color="#3498db"))
    fig.add_trace(go.Bar(name=pc, x=top["Account"], y=top[pc], marker_color="#00d4aa"))
    fig.update_layout(title=f"Top 15 Accounts - {pc} vs {cc}",
                      barmode="group", xaxis_tickangle=-45, height=480,
                      legend=dict(orientation="h", y=1.02, x=1, xanchor="right"))
    return fig

def create_rep_chart(df):
    import plotly.express as px
    pc = df["_primary_col"].iloc[0]
    d  = df.groupby("Sales Rep").agg(Sales=(pc,"sum")).reset_index()
    d  = d[d["Sales Rep"] != ""].sort_values("Sales", ascending=True)
    if d.empty: return None
    return px.bar(d, y="Sales Rep", x="Sales", orientation="h",
                  title=f"{pc} by Rep", color="Sales",
                  color_continuous_scale="Blues", height=380)

def create_tier_chart(df):
    import plotly.express as px
    pc = df["_primary_col"].iloc[0]
    d  = df.groupby("Tier").agg(Sales=(pc,"sum")).reset_index()
    d  = d[d["Tier"] != ""]
    if d.empty: return None
    return px.pie(d, values="Sales", names="Tier", title=f"{pc} by Tier",
                  color_discrete_sequence=px.colors.qualitative.Set2, height=380)

def create_scatter_chart(df):
    import plotly.express as px
    pc = df["_primary_col"].iloc[0]; cc = df["_comparison_col"].iloc[0]
    if cc not in df.columns: return None
    plot_df = df[(df[pc]>0)|(df[cc]>0)].copy()
    if plot_df.empty: return None
    fig = px.scatter(plot_df, x=cc, y=pc, size=pc,
                     color="Tier" if plot_df["Tier"].any() else None,
                     hover_name="Account",
                     hover_data={"$ Change":True,"% Change":True},
                     title=f"{pc} vs {cc}", height=460)
    mx = max(plot_df[cc].max(), plot_df[pc].max())
    fig.add_shape(type="line", x0=0, y0=0, x1=mx, y1=mx,
                  line=dict(color="gray", dash="dash"))
    return fig

# =============================================================================
# PERIOD RESOLUTION
# =============================================================================

def _quarter_bounds(year, q):
    starts = {1:(1,1), 2:(4,1), 3:(7,1), 4:(10,1)}
    ends   = {1:(3,31),2:(6,30),3:(9,30),4:(12,31)}
    s,e = starts[q],ends[q]
    return datetime(year,s[0],s[1]).date(), datetime(year,e[0],e[1]).date()

def resolve_primary_period(name, today, cy, cm, cq, cs=None, ce=None):
    import calendar
    if name == "This Month":
        ld = calendar.monthrange(cy,cm)[1]
        return datetime(cy,cm,1).strftime("%b %Y"), datetime(cy,cm,1).date(), min(datetime(cy,cm,ld).date(),today)
    if name == "Last Month":
        lm = cm-1 if cm>1 else 12; ly = cy if cm>1 else cy-1
        ld = calendar.monthrange(ly,lm)[1]
        return datetime(ly,lm,1).strftime("%b %Y"), datetime(ly,lm,1).date(), datetime(ly,lm,ld).date()
    if name == "This Quarter":
        s,e = _quarter_bounds(cy,cq)
        return f"Q{cq} {cy}", s, min(e,today)
    if name == "Last Quarter":
        lq = cq-1 if cq>1 else 4; ly = cy if cq>1 else cy-1
        s,e = _quarter_bounds(ly,lq)
        return f"Q{lq} {ly}", s, e
    if name == "Year to Date":
        return f"{cy} YTD", datetime(cy,1,1).date(), today
    if name == "Last 12 Months":
        return "Last 12 Months", today.replace(year=today.year-1), today
    if name == "This Year (Full)":
        return str(cy), datetime(cy,1,1).date(), datetime(cy,12,31).date()
    if name == "Last Year (Full)":
        return str(cy-1), datetime(cy-1,1,1).date(), datetime(cy-1,12,31).date()
    if name == "Last 30 Days":
        return "Last 30 Days", today-timedelta(days=30), today
    if name == "Last 60 Days":
        return "Last 60 Days", today-timedelta(days=60), today
    if name == "Last 90 Days":
        return "Last 90 Days", today-timedelta(days=90), today
    if name == "Custom Range":
        s,e = cs or today, ce or today
        return f"{s.strftime('%b %d')} - {e.strftime('%b %d, %Y')}", s, e
    return f"{cy} YTD", datetime(cy,1,1).date(), today

def same_period_prior_year(start, end):
    try:
        cs = start.replace(year=start.year-1)
    except ValueError:
        cs = start - timedelta(days=365)
    try:
        ce = end.replace(year=end.year-1)
    except ValueError:
        ce = end - timedelta(days=365)
    return cs, ce

# =============================================================================
# MAIN
# =============================================================================

def main():
    inject_css()

    can_generate = False

    st.markdown(f"""
    <div class="main-header">
        <h1>📊 {BRANDING['company_name']} Management Reports</h1>
        <p style="color:#666;margin:0.2rem 0 0;">Sales Intelligence Dashboard</p>
    </div>
    """, unsafe_allow_html=True)

    # =========================================================================
    # SIDEBAR
    # =========================================================================
    with st.sidebar:
        st.header("⚙️ Configuration")

        st.subheader("📦 Cin7 API")
        cin7_user = st.text_input("Username", value=get_secret("CIN7_USERNAME"), key="cin7_user")
        cin7_key  = st.text_input("API Key",  value=get_secret("CIN7_API_KEY"),
                                   type="password", key="cin7_key")

        can_generate = bool(cin7_user and cin7_key)

        if cin7_user and cin7_key:
            if st.button("Test Cin7", key="test_cin7"):
                ok, msg = test_cin7_connection(cin7_user, cin7_key)
                (st.success if ok else st.error)(f"{'OK' if ok else 'ERR'}: {msg}")

        st.divider()

        st.subheader("🟠 HubSpot API")
        hubspot_key   = st.text_input("Private App Token", value=get_secret("HUBSPOT_API_KEY"),
                                       type="password", key="hubspot_key")
        tier_property = st.text_input("Tier Property",
                                       value=get_secret("HUBSPOT_TIER_PROPERTY","commission_tier"),
                                       key="tier_prop")

        if hubspot_key:
            if st.button("Test HubSpot", key="test_hs"):
                ok, msg = test_hubspot_connection(hubspot_key)
                (st.success if ok else st.error)(f"{'OK' if ok else 'ERR'}: {msg}")

        st.divider()

        st.subheader("📅 Report Period")
        today = datetime.now().date()
        cy = today.year; cm = today.month; cq = (cm-1)//3+1

        PERIOD_OPTIONS = [
            "This Month","Last Month","This Quarter","Last Quarter",
            "Year to Date","Last 12 Months","This Year (Full)","Last Year (Full)",
            "Last 30 Days","Last 60 Days","Last 90 Days","Custom Range",
        ]
        cfg          = st.session_state.config_loaded
        saved_period = cfg.get("last_period", "Year to Date")
        period_idx   = PERIOD_OPTIONS.index(saved_period) if saved_period in PERIOD_OPTIONS else 4
        selected_period = st.selectbox("Primary Period", PERIOD_OPTIONS,
                                        index=period_idx, key="primary_period")

        custom_start = custom_end = None
        if selected_period == "Custom Range":
            custom_start = st.date_input("From", value=datetime(cy,1,1).date(), key="custom_start")
            custom_end   = st.date_input("To",   value=today,                   key="custom_end")

        p_label, p_start, p_end = resolve_primary_period(
            selected_period, today, cy, cm, cq, custom_start, custom_end)

        COMPARE_OPTIONS = ["Same Period Last Year","Previous Period",
                           "Custom Comparison Range","None"]
        saved_compare = cfg.get("last_compare","Same Period Last Year")
        compare_idx   = COMPARE_OPTIONS.index(saved_compare) if saved_compare in COMPARE_OPTIONS else 0
        compare_to    = st.selectbox("Compare Against", COMPARE_OPTIONS,
                                      index=compare_idx, key="compare_to")

        comp_cs = comp_ce = None
        if compare_to == "Custom Comparison Range":
            try:    dcs = p_start.replace(year=p_start.year-1)
            except: dcs = p_start - timedelta(days=365)
            try:    dce = p_end.replace(year=p_end.year-1)
            except: dce = p_end - timedelta(days=365)
            comp_cs = st.date_input("Compare From", value=dcs, key="comp_cstart")
            comp_ce = st.date_input("Compare To",   value=dce, key="comp_cend")

        comp = None
        if compare_to == "Same Period Last Year":
            cs, ce = same_period_prior_year(p_start, p_end)
            c_label = p_label.replace(str(p_start.year), str(cs.year)) \
                      if str(p_start.year) in p_label else f"{p_label} (Prior Year)"
            comp = (c_label, cs, ce)
        elif compare_to == "Previous Period":
            delta = p_end - p_start
            ce = p_start - timedelta(days=1); cs = ce - delta
            comp = (f"{cs.strftime('%b %d')} - {ce.strftime('%b %d, %Y')}", cs, ce)
        elif compare_to == "Custom Comparison Range" and comp_cs and comp_ce:
            comp = (f"{comp_cs.strftime('%b %d')} - {comp_ce.strftime('%b %d, %Y')}", comp_cs, comp_ce)

        if comp:
            periods = [{"label":comp[0],"start":comp[1],"end":comp[2]},
                       {"label":p_label,"start":p_start,"end":p_end}]
        else:
            periods = [{"label":p_label,"start":p_start,"end":p_end}]

        st.caption(f"📅 **{p_label}:** {p_start.strftime('%b %d, %Y')} to {p_end.strftime('%b %d, %Y')}")
        if comp:
            pd_ = (p_end-p_start).days+1; cd_ = (comp[2]-comp[1]).days+1
            st.caption(f"📅 **vs. {comp[0]}:** {comp[1].strftime('%b %d, %Y')} to {comp[2].strftime('%b %d, %Y')}")
            st.caption(f"{'OK' if pd_==cd_ else 'NOTE'}: {pd_}d vs {cd_}d")

        st.divider()

        st.subheader("🚫 Excluded Domains")
        _exc_raw = load_config().get("excluded_domains",
                                      get_secret("EXCLUDED_DOMAINS","vivantskincare.com"))
        excluded_input = st.text_area("One domain per line",
            value="\n".join(d.strip() for d in _exc_raw.split(",") if d.strip()),
            height=70, key="excluded_domains_input")
        if st.button("💾 Save Exclusions", use_container_width=True):
            save_config({"excluded_domains": ", ".join(
                d.strip().lower() for d in excluded_input.splitlines() if d.strip())})
            cache_clear_all()
            st.session_state.report_data = None
            st.success("Saved.")
            st.rerun()

        meta           = _load_cache_meta()
        cached_periods = [v["label"] for k,v in meta.items() if k.startswith("cin7_")]
        hs_entry       = meta.get("hubspot")
        contacts_entry = meta.get("contacts")
        if cached_periods:
            parts = [f"{len(cached_periods)} order period(s)"]
            if hs_entry:
                h = (datetime.now()-datetime.fromisoformat(hs_entry["saved_at"])).total_seconds()/3600
                parts.append(f"HubSpot {h:.0f}h")
            if contacts_entry:
                h = (datetime.now()-datetime.fromisoformat(contacts_entry["saved_at"])).total_seconds()/3600
                parts.append(f"Contacts {h:.0f}h")
            st.caption("💾 " + " · ".join(parts))
            raw_csv = export_raw_orders_csv()
            if raw_csv:
                st.download_button(
                    "⬇️ Download Raw Cin7 Data",
                    data=raw_csv,
                    file_name=f"cin7_raw_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    help="All cached orders, sorted by date — one row per order"
                )
            if st.button("🗑️ Clear Cache", use_container_width=True):
                cache_clear_all()
                st.session_state.report_data = None
                st.session_state.cin7_customers = {}
                st.rerun()

        st.divider()

        # ── PRELOAD: bulk-cache all months in a selectable range ─────────────
        st.subheader("📥 Preload Historical Data")
        today_yr = datetime.now().year
        preload_years = st.slider(
            "Years of history to load",
            min_value=1, max_value=5, value=2,
            help="Loads every month from Jan of that year to today and caches it. "
                 "Subsequent Generate clicks will be instant."
        )
        start_preload_year = today_yr - preload_years + 1

        col_pre1, col_pre2 = st.columns(2)
        with col_pre1:
            st.caption(f"From: Jan {start_preload_year}")
        with col_pre2:
            st.caption(f"To: {datetime.now().strftime('%b %Y')}")

        if st.button("📥 Preload Data", use_container_width=True, disabled=not can_generate):
            prog = st.progress(0, text="Starting…")
            summary = preload_months(
                cin7_user, cin7_key,
                start_preload_year, 1,
                datetime.now().year, datetime.now().month,
                prog
            )
            st.success(
                f"✅ Preload complete — {summary['orders']:,} orders cached "
                f"({summary['fetched']} months fetched, {summary['skipped']} skipped)"
            )
            st.rerun()

        st.divider()

        # ── AUTO-FETCH: trigger on load if secrets configured & no report yet ──
        auto_fetch = (can_generate and
                      st.session_state.report_data is None and
                      all(cache_has_orders(p["label"]) for p in periods))

        if st.button("🔄 Generate Report", type="primary",
                     use_container_width=True, disabled=not can_generate) or auto_fetch:

            save_config({"last_period": selected_period, "last_compare": compare_to})
            st.session_state.periods = periods

            has_all_cache = all(cache_has_orders(p["label"]) for p in periods)

            if has_all_cache and not auto_fetch:
                quick_orders = []
                for p in periods:
                    cached = cache_load_orders_any(p["label"])
                    if cached:
                        quick_orders.extend(cached)

                if quick_orders:
                    hs_tiers, hs_owners = cache_load_hubspot()
                    hs_tiers  = hs_tiers  or {}
                    hs_owners = hs_owners or {}
                    customers = cache_load_contacts() or {}
                    staff     = st.session_state.cin7_staff or {}

                    cdata, audit = aggregate_orders_by_company(
                        quick_orders, periods, cin7_staff=staff)
                    df_quick = build_report_dataframe(
                        cdata, hs_tiers, periods, hs_owners, {},
                        cin7_customers=customers)

                    if not df_quick.empty:
                        st.session_state.report_data = df_quick
                        st.session_state.audit       = audit
                        st.session_state.periods     = periods

            with st.spinner("Refreshing..."):
                result = run_full_fetch(cin7_user, cin7_key, hubspot_key,
                                        tier_property, periods)
                st.session_state.report_data = result["df"]
                st.session_state.audit       = result["audit"]
                st.session_state.periods     = periods
            st.rerun()

        elif auto_fetch:
            quick_orders = []
            for p in periods:
                cached = cache_load_orders_any(p["label"])
                if cached:
                    quick_orders.extend(cached)
            if quick_orders:
                hs_tiers, hs_owners = cache_load_hubspot()
                customers = cache_load_contacts() or {}
                staff     = st.session_state.cin7_staff or {}
                cdata, audit = aggregate_orders_by_company(
                    quick_orders, periods, cin7_staff=staff)
                df_auto = build_report_dataframe(
                    cdata, hs_tiers or {}, periods, hs_owners or {}, {},
                    cin7_customers=customers)
                if not df_auto.empty:
                    st.session_state.report_data = df_auto
                    st.session_state.audit       = audit
                    st.session_state.periods     = periods
                    st.rerun()

    # =========================================================================
    # MAIN CONTENT
    # =========================================================================
    df = st.session_state.report_data

    REQUIRED = {"Account","$ Change","% Change","Tier","Sales Rep","_primary_col","_comparison_col"}
    if df is not None and not REQUIRED.issubset(df.columns):
        st.session_state.report_data = None
        df = None

    if df is None:
        debug = st.session_state.get("_debug")
        if debug:
            st.subheader("🐛 Debug — Last Fetch Results")
            st.json(debug)
            api_debug = st.session_state.get("_api_debug")
            if api_debug:
                st.subheader("🌐 Cin7 API Response")
                st.json(api_debug)
            # Also try reading from file
            try:
                import json as _json
                with open(".cin7_api_debug.json") as _f:
                    file_debug = _json.load(_f)
                st.subheader("🌐 Cin7 API Response (from file)")
                st.json(file_debug)
            except Exception:
                st.warning("No API debug file found yet — fetch may not have reached the API call.")
            st.divider()

        audit = st.session_state.get("audit")
        if audit and audit.get("total_raw", 0) > 0:
            st.warning("Report returned no qualifying wholesale rows. See breakdown below.")
            total = audit["total_raw"]
            st.write(f"**Raw orders from Cin7:** {total:,}")
            st.write(f"**Excluded Shopify/Retail:** {audit['excluded_source']:,}")
            st.write(f"**Excluded domain:** {audit.get('excluded_domain', 0):,}")
            st.write(f"**Outside date range:** {audit['excluded_no_period']:,}")
            st.write(f"**$0 total:** {audit['excluded_zero_total']:,}")
            st.write(f"**No company name:** {audit['unknown_company']:,}")
            if audit.get("excluded_sources"):
                st.write("**Source values seen:**", audit["excluded_sources"])
            if audit.get("sample_excluded"):
                st.write("**Sample excluded orders:**")
                st.dataframe(pd.DataFrame(audit["sample_excluded"]), use_container_width=True)
        else:
            st.info("👈 Configure your API credentials and click **Generate Report** to begin.")
            st.markdown("""
            | Column | Description |
            |---|---|
            | **Account** | Company / account name |
            | **YTD Sales** | Revenue for selected primary period |
            | **Prior Year** | Revenue for comparison period |
            | **$ Change** | Dollar difference |
            | **% Change** | Growth / decline % |
            | **Tier** | Commission tier (HA, 10%, 6%) |
            | **Sales Rep** | Assigned rep from approved account list |
            """)
        return

    primary_col = df["_primary_col"].iloc[0]
    comp_col    = df["_comparison_col"].iloc[0]
    periods     = st.session_state.get("periods", [])

    # ── Filters ───────────────────────────────────────────────────────────────
    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        reps = ["All"] + sorted([r for r in df["Sales Rep"].unique() if r])
        selected_rep = st.selectbox("Sales Rep", reps)
    with fc2:
        tiers_list = sorted([t for t in df["Tier"].unique() if t])
        has_no_tier = df["Tier"].eq("").any() or df["Tier"].isna().any()
        selected_tier = st.selectbox("Tier", ["All"]+tiers_list+(["(No Tier)"] if has_no_tier else []))
    with fc3:
        min_sales = st.number_input(f"Min {primary_col} ($)", min_value=0, value=0, step=500)
    with fc4:
        sort_options = [
            f"{primary_col} down", f"{primary_col} up",
            f"{comp_col} down", f"{comp_col} up",
            "$ Change down", "$ Change up",
            "% Change down", "% Change up",
            "Account A-Z",
        ]
        sort_by = st.selectbox("Sort By", sort_options)

    fdf = df.copy()
    if selected_rep  != "All": fdf = fdf[fdf["Sales Rep"]==selected_rep]
    if selected_tier == "(No Tier)": fdf = fdf[fdf["Tier"].eq("")|fdf["Tier"].isna()]
    elif selected_tier != "All":     fdf = fdf[fdf["Tier"]==selected_tier]
    if min_sales > 0: fdf = fdf[fdf[primary_col]>=min_sales]

    sort_map = {
        f"{primary_col} down":(primary_col,False), f"{primary_col} up":(primary_col,True),
        f"{comp_col} down":(comp_col,False),        f"{comp_col} up":(comp_col,True),
        "$ Change down":("$ Change",False),          "$ Change up":("$ Change",True),
        "% Change down":("% Change",False),          "% Change up":("% Change",True),
        "Account A-Z":("Account",True),
    }
    scol, sasc = sort_map.get(sort_by, (primary_col, False))
    if scol in fdf.columns:
        fdf = fdf.sort_values(scol, ascending=sasc).reset_index(drop=True)

    # ── Summary metrics ────────────────────────────────────────────────────────
    st.divider()
    tp = fdf[primary_col].sum()
    tc = fdf[comp_col].sum() if comp_col in fdf.columns else 0
    td = fdf["$ Change"].sum()
    tp_pct = ((tp-tc)/tc*100) if tc > 0 else 0
    growing   = (fdf["% Change"]>0).sum()
    declining = (fdf["% Change"]<0).sum()

    m1,m2,m3,m4,m5,m6 = st.columns(6)
    with m1: st.metric("Accounts",  f"{len(fdf):,}")
    with m2: st.metric(primary_col, f"${tp:,.0f}")
    with m3: st.metric(comp_col,    f"${tc:,.0f}")
    with m4: st.metric("$ Change",  f"${td:+,.0f}")
    with m5: st.metric("% Change",  f"{tp_pct:+.1f}%")
    with m6: st.metric("Up / Down", f"{growing} / {declining}")

    st.divider()

    tab1, tab2, tab3, tab4 = st.tabs(
        ["📋 Account Report","📈 Charts","📤 Export","🔍 Data Audit"])

    # TAB 1
    with tab1:
        if len(periods)==2:
            pri,cmp = periods[1],periods[0]
            st.markdown(
                f"**{pri['label']}** {pri['start'].strftime('%b %d, %Y')} to {pri['end'].strftime('%b %d, %Y')}"
                f"  |  **vs. {cmp['label']}** {cmp['start'].strftime('%b %d, %Y')} to {cmp['end'].strftime('%b %d, %Y')}")
        elif len(periods)==1:
            pri = periods[0]
            st.markdown(f"**{pri['label']}** {pri['start'].strftime('%b %d, %Y')} to {pri['end'].strftime('%b %d, %Y')}")
        st.caption(f"{len(fdf)} accounts")

        display_cols = [c for c in ["Account",primary_col,comp_col,"$ Change","% Change",
                                     "Tier","Sales Rep"] if c in fdf.columns]
        col_cfg = {
            "Account":   st.column_config.TextColumn("Account", width="medium"),
            primary_col: st.column_config.NumberColumn(primary_col, format="$%.2f", width="small"),
            "$ Change":  st.column_config.NumberColumn("$ Change",  format="$%.2f", width="small"),
            "% Change":  st.column_config.NumberColumn("% Change",  format="%.1f%%", width="small"),
            "Tier":      st.column_config.TextColumn("Tier",       width="small"),
            "Sales Rep": st.column_config.TextColumn("Sales Rep",  width="small"),
        }
        if comp_col in fdf.columns:
            col_cfg[comp_col] = st.column_config.NumberColumn(comp_col, format="$%.2f", width="small")

        st.dataframe(fdf[display_cols], use_container_width=True, hide_index=True,
                     column_config=col_cfg,
                     height=min(600, (len(fdf)+1)*35+38))

    # TAB 2
    with tab2:
        st.subheader("📈 Visual Analytics")
        if not fdf.empty:
            st.plotly_chart(create_yoy_chart(fdf), use_container_width=True)
        cc1,cc2 = st.columns(2)
        with cc1:
            fig = create_rep_chart(fdf)
            if fig: st.plotly_chart(fig, use_container_width=True)
        with cc2:
            fig = create_tier_chart(fdf)
            if fig: st.plotly_chart(fig, use_container_width=True)
        fig = create_scatter_chart(fdf)
        if fig:
            st.plotly_chart(fig, use_container_width=True)
            st.caption("Accounts above the diagonal are growing; below are declining.")

    # TAB 3
    with tab3:
        st.subheader("📤 Export Report")
        export_cols = [c for c in ["Account",primary_col,comp_col,"$ Change","% Change",
                                    "Tier","Sales Rep"] if c in fdf.columns]
        export_df = fdf[export_cols].copy()
        ec1,ec2 = st.columns(2)
        with ec1:
            st.markdown("### Excel")
            st.write(f"{len(export_df)} accounts")
            st.download_button("Download Excel", data=export_to_excel(export_df),
                file_name=f"sales_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with ec2:
            st.markdown("### CSV")
            st.write(f"{len(export_df)} accounts")
            st.download_button("Download CSV", data=export_df.to_csv(index=False),
                file_name=f"sales_report_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv")

    # TAB 4
    with tab4:
        audit = st.session_state.get("audit")
        if not audit:
            st.info("Run a report to see the data audit.")
        else:
            st.subheader("🔍 Data Audit")
            total = audit["total_raw"]; kept = audit["included"]; dropped = total-kept
            a1,a2,a3,a4 = st.columns(4)
            with a1: st.metric("Raw Orders",       f"{total:,}")
            with a2: st.metric("Included",         f"{kept:,}",
                                delta=f"{kept/total*100:.1f}%" if total else "0%")
            with a3: st.metric("Excluded",         f"{dropped:,}",
                                delta=f"-{dropped/total*100:.1f}%" if total else "0%",
                                delta_color="inverse")
            with a4: st.metric("Unique Companies", f"{audit.get('unique_companies',0):,}")

            if "Tier" in df.columns:
                untiered = df["Tier"].eq("").sum() + df["Tier"].isna().sum()
                if untiered:
                    st.warning(f"{untiered} accounts have no tier assigned.")

            st.divider()
            st.markdown("#### Exclusion Breakdown")
            excl_df = pd.DataFrame({
                "Reason":["Shopify/Retail (B2C)","Excluded email domain","Outside date windows",
                           "$0 total","No company name"],
                "Count": [audit["excluded_source"], audit.get("excluded_domain",0),
                           audit["excluded_no_period"], audit["excluded_zero_total"],
                           audit["unknown_company"]],
            })
            excl_df["% of Raw"] = excl_df["Count"].apply(
                lambda x: f"{x/total*100:.1f}%" if total else "0%")
            st.dataframe(excl_df, use_container_width=True, hide_index=True)

            if audit.get("excluded_domain_counts"):
                st.markdown("#### Excluded by Domain")
                st.dataframe(pd.DataFrame([{"Domain":k,"Dropped":v} for k,v in
                    sorted(audit["excluded_domain_counts"].items(),key=lambda x:-x[1])]),
                    use_container_width=True, hide_index=True)

            if audit["excluded_sources"]:
                st.markdown("#### Excluded Source Values")
                st.dataframe(pd.DataFrame([{"Source":k,"Dropped":v} for k,v in
                    sorted(audit["excluded_sources"].items(),key=lambda x:-x[1])]),
                    use_container_width=True, hide_index=True)

            st.divider()
            st.markdown("#### By Period")
            st.dataframe(pd.DataFrame([
                {"Period":lbl,"Included":s["included"],
                 "Revenue":f"${s['revenue']:,.2f}","Excluded B2C":s["excluded_source"]}
                for lbl,s in audit["by_period"].items()
            ]), use_container_width=True, hide_index=True)

            if audit["sample_excluded"]:
                st.divider()
                st.markdown("#### Sample Excluded Orders")
                st.dataframe(pd.DataFrame(audit["sample_excluded"]),
                             use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown(
        f"<p style='text-align:center;color:#888;font-size:0.75rem;'>"
        f"Powered by {BRANDING['company_name']}</p>",
        unsafe_allow_html=True)


# =============================================================================
# ENTRY POINT
# =============================================================================
main()
